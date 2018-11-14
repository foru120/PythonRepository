from openpyxl import load_workbook
from wordcloud import WordCloud, STOPWORDS
import matplotlib.pyplot as plt

def create_word_cloud(data_dict):
    wordcloud = WordCloud(font_path='/home/kyh/dataset/OdibeeSans-Regular.ttf',
                          stopwords=STOPWORDS,
                          background_color='black',
                          width=500,
                          height=400,
                          colormap='spring').fit_words(data_dict)
    return wordcloud

def save_image(word_cloud, name):
    fig = plt.figure(figsize=(18, 9))
    plt.title(name)
    plt.imshow(word_cloud)
    plt.axis('off')
    fig.savefig('/home/kyh/PycharmProjects/PythonRepository/Projects/AnalysisProject/wordcloud/'
                + name + '_wordcloud.jpg', dpi=100)

review_excel = load_workbook(filename='/home/kyh/dataset/wordcloud_grad_con.xlsx')

for sheet_name in review_excel.sheetnames:
    if sheet_name == 'edit_rank_word':  # edit_rank_word 엑셀 시트
        neg_edit_dict = dict()
        pos_edit_dict = dict()
        cnt = 0
        for row in review_excel[sheet_name].rows:
            cnt += 1
            if cnt == 1: continue  # 헤더 제거

            if row[0].value is not None:  # neg_word
                neg_edit_dict[row[0].value] = int(row[1].value)

            if row[2].value is not None:  # pos_word
                pos_edit_dict[row[2].value] = int(row[3].value)

        neg_word_cloud = create_word_cloud(neg_edit_dict)
        pos_word_cloud = create_word_cloud(pos_edit_dict)

        save_image(neg_word_cloud, 'neg_edit_rank')
        save_image(pos_word_cloud, 'pos_edit_rank')
    elif sheet_name == 'pos_Nrank1':  # pos_Nrank1 엑셀 시트
        small_ind_dict = dict()
        small_col_dict = dict()
        large_ind_dict = dict()
        large_col_dict = dict()

        cnt = 0
        for row in review_excel[sheet_name].rows:
            cnt += 1
            if cnt == 1: continue  # 헤더 제거

            if row[0].value is not None:  # small power distance individualist
                small_ind_dict[row[0].value] = int(row[1].value)

            if row[3].value is not None:  # small power distance collectivist
                small_col_dict[row[3].value] = int(row[4].value)

            if row[6].value is not None:  # large power distance individualist
                large_ind_dict[row[6].value] = int(row[7].value)

            if row[9].value is not None:  # large power distance collectivist
                large_col_dict[row[9].value] = int(row[10].value)

        small_ind_cloud = create_word_cloud(small_ind_dict)
        small_col_cloud = create_word_cloud(small_col_dict)
        large_ind_cloud = create_word_cloud(large_ind_dict)
        large_col_cloud = create_word_cloud(large_col_dict)

        save_image(small_ind_cloud, 'Positive Small power distance Individualist')
        save_image(small_col_cloud, 'Positive Small power distance Collectivist')
        save_image(large_ind_cloud, 'Positive Large power distance Individualist')
        save_image(large_col_cloud, 'Positive Large power distance Collectivist')
    elif sheet_name == 'neg_Nrank':
        small_ind_dict = dict()
        small_col_dict = dict()
        large_ind_dict = dict()
        large_col_dict = dict()

        cnt = 0
        for row in review_excel[sheet_name].rows:
            cnt += 1
            if cnt == 1: continue  # 헤더 제거

            if row[0].value is not None:  # small power distance individualist
                small_ind_dict[row[0].value] = int(row[1].value)

            if row[3].value is not None:  # small power distance collectivist
                small_col_dict[row[3].value] = int(row[4].value)

            if row[6].value is not None:  # large power distance individualist
                large_ind_dict[row[6].value] = int(row[7].value)

            if row[9].value is not None:  # large power distance collectivist
                large_col_dict[row[9].value] = int(row[10].value)

        small_ind_cloud = create_word_cloud(small_ind_dict)
        small_col_cloud = create_word_cloud(small_col_dict)
        large_ind_cloud = create_word_cloud(large_ind_dict)
        large_col_cloud = create_word_cloud(large_col_dict)

        save_image(small_ind_cloud, 'Negative Small power distance Individualist')
        save_image(small_col_cloud, 'Negative Small power distance Collectivist')
        save_image(large_ind_cloud, 'Negative Large power distance Individualist')
        save_image(large_col_cloud, 'Negative Large power distance Collectivist')